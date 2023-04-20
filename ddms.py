import csv
import datetime
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.styles import PatternFill, Font, Alignment

def open_file_dialog():
    Tk().withdraw()
    file_path = filedialog.askopenfilename(title="Select the input file", filetypes=[("Text Files", "*.txt")])
    return file_path

def remove_illegal_chars(cell_value):
    return ''.join(c for c in cell_value if c.isprintable())

now = datetime.datetime.now()
date_string = now.strftime('%Y-%m-%d')
file_name = f'Pick_List_{date_string}.xlsx'

input_file = open_file_dialog()

# Read the data from the text file
data = []
with open(input_file, 'r') as f:
    reader = csv.reader(f, delimiter='\t')
    for i, row in enumerate(reader):
        if i < 2:
            continue
        new_row = []
        for cell in row:
            new_row.extend(cell.split('|'))
        data.append(new_row)

# Create a new Excel workbook and add the data
wb = Workbook()
ws = wb.active

# Set the headers
headers = data.pop(0)
header_font = Font(bold=True)
header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Add the remaining data to the worksheet
for row_num, row in enumerate(data, start=2):
    cleaned_row = [remove_illegal_chars(cell) for cell in row]
    for col_num, cell_value in enumerate(cleaned_row, start=1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Delete columns that are not needed
ws.delete_cols(2, 1)
ws.delete_cols(2, 1)
ws.delete_cols(14, 1)
ws.delete_cols(17, 4)
ws.delete_cols(20, 7)

# Save the workbook as an .xlsx file
wb.save(file_name)
